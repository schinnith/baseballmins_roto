import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# === CONFIG ===
CSV_PATH = r"C:\Users\chris\nfbc_projects\Python Projects\espn_roto_standings\espn_roto_standings.csv"
OUTPUT_XLSX = CSV_PATH.replace(".csv", "_formatted.xlsx")
ASCENDING_CATEGORIES = ['ERA', 'WHIP']
DESCENDING_CATEGORIES = ['R', 'HR', 'RBI', 'SB', 'AVG', 'K', 'W', 'SV']
CATEGORIES = DESCENDING_CATEGORIES + ASCENDING_CATEGORIES

# === ROTO POINTS ===
def calculate_roto_points(series, ascending):
    ranked = series.rank(ascending=ascending, method='min')
    points = ranked.copy()
    value_counts = ranked.value_counts()
    for rank, count in value_counts.items():
        total = sum(11 - r for r in range(int(rank), int(rank + count)))
        points[ranked == rank] = total / count
    return points

# === FORMATTING ===
def apply_conditional_formatting(ws, point_cols, total_col, num_teams):
    for col_letter in point_cols:
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{num_teams+1}",
            ColorScaleRule(start_type='min', start_color='FFAAAA',
                           mid_type='percentile', mid_value=50, mid_color='FFFFAA',
                           end_type='max', end_color='AAFFAA')
        )
    ws.conditional_formatting.add(
        f"{total_col}2:{total_col}{num_teams+1}",
        ColorScaleRule(start_type='min', start_color='FEE1E1',
                       mid_type='percentile', mid_value=50, mid_color='FFFFBB',
                       end_type='max', end_color='B7F7B7')
    )

def bold_first_place(ws):
    for cell in ws["A2:Z2"][0]:
        cell.font = Font(bold=True)

def add_table_formatting(ws, name, df_shape):
    num_rows, num_cols = df_shape
    end_col_letter = get_column_letter(num_cols)
    ref = f"A1:{end_col_letter}{num_rows+1}"  # +1 to include header
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

def autofit_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

# === READ FROM STRUCTURED ESPN CSV ===
def parse_structured_csv(filepath):
    # Read as a table, skip empty lines
    df = pd.read_csv(filepath, header=None, encoding='latin1')
    # Find the row with 'R' as the first column (stat headers)
    stat_header_row = df[df[0] == 'R'].index[0]
    # The next row is the first team stats row
    stats_start = stat_header_row + 1

    # Find the first empty row after stats (end of stats)
    stats_end = stats_start
    while stats_end < len(df) and pd.notnull(df.iloc[stats_end, 0]):
        stats_end += 1

    # Get team names from the rows above the stat header
    team_rows = df.loc[3:12, 1].tolist()  # 10 teams, adjust if needed

    # Get stat headers
    stat_headers = df.iloc[stat_header_row].tolist()
    # Get stats
    stats = df.iloc[stats_start:stats_end, :len(stat_headers)]
    stats.columns = stat_headers
    stats = stats.reset_index(drop=True)

    # Insert team names
    stats.insert(0, 'Team', team_rows)
    return stats

# === MAIN ===
def main():
    df = parse_structured_csv(CSV_PATH)
    roto = df[['Team']].copy()

    for cat in DESCENDING_CATEGORIES:
        roto[f"{cat}_Pts"] = calculate_roto_points(df[cat], ascending=False)
    for cat in ASCENDING_CATEGORIES:
        roto[f"{cat}_Pts"] = calculate_roto_points(df[cat], ascending=True)

    point_cols = [col for col in roto.columns if col.endswith('_Pts')]
    roto['Total_Points'] = roto[point_cols].sum(axis=1)
    roto_sorted = roto.sort_values(by='Total_Points', ascending=False).reset_index(drop=True)

    with pd.ExcelWriter(OUTPUT_XLSX, engine='openpyxl') as writer:
        roto_sorted.to_excel(writer, sheet_name='Roto Standings', index=False)
        df.to_excel(writer, sheet_name='Raw Totals', index=False)

    wb = load_workbook(OUTPUT_XLSX)
    ws1 = wb['Roto Standings']
    ws2 = wb['Raw Totals']
    add_table_formatting(ws1, "RotoPoints", roto_sorted.shape)
    add_table_formatting(ws2, "RawStats", df.shape)

    point_letters = [get_column_letter(i+2) for i in range(len(point_cols))]  # skip 'Team' column
    total_col = get_column_letter(len(point_cols) + 2)

    apply_conditional_formatting(ws1, point_letters, total_col, len(roto_sorted))
    bold_first_place(ws1)
    autofit_columns(ws1)
    autofit_columns(ws2)
    wb.save(OUTPUT_XLSX)

    print(f"\n✅ Excel file saved: {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()
